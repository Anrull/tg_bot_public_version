import openpyxl
import json
import sys

# 19 2


def update(path_of_xlsx):
    # sys.stdout = open("dicts/teachers.py", "w", encoding="UTF-8")
    book = openpyxl.load_workbook(path_of_xlsx, read_only=True, data_only=True)

    MEGA_dict = {}
    HUGE_dict = {}
    mega_count1 = 0

    for k in range(1, 3):
        book.active = k
        sheet = book.active

        # mega_list = ["6A", "6B", "7A", "7B", "7C", "8A", "8B", "9A", "9B", "10A", "10B", "11A", "11B"]
        mega_dict = {}
        
        # print(sheet[1][0].value)
        
        # title = sheet[2][19].value.lower()

        for i in range(5, sheet.max_row + 1):
            dict_object = {}
            dict_classes_day = {}
            lst = []
            count = 0
            # name = sheet[i][j].value
            name = sheet[i][1].value
            # print(name)
            
            for j in range(2, sheet.max_column):
                num = sheet[4][j].value
                classes = sheet[i][j].value
                day = "Понедельник"
                if sheet[3][j].value:
                    day = sheet[3][j].value
                if str(num) != "9":
                    lst.append([classes, num])
                    if j == sheet.max_column:
                        # dict_classes_day[day] = lst
                        # dict_object[count] = dict_classes_day
                        dict_object[count] = lst
                        lst = []
                        count += 1
                        dict_classes_day = {}
                elif j == sheet.max_column:
                    # dict_classes_day[day] = lst
                    # dict_object[count] = dict_classes_day
                    dict_object[count] = lst
                    lst = []
                    count += 1
                    dict_classes_day = {}
                else:
                    # dict_classes_day[day] = lst
                    # dict_object[count] = dict_classes_day
                    dict_object[count] = lst
                    lst = []
                    count += 1
                    dict_classes_day = {}
            
            mega_dict[name] = dict_object
            # mega_count += 1
            count = 0
        
        MEGA_dict[mega_count1] = mega_dict
        mega_dict = {}
        mega_count1 += 1
        
    # print(MEGA_dict)
    
    # print(MEGA_dict[1])


    # print(MEGA_dict)
    json_dict = json.dumps(obj=MEGA_dict)

    sys.stdout = open("dicts/teachers.json", "w")
    print(json_dict)


if __name__ == "__main__":
    update("Books/raspisanie_23-24_gotovo_2_polugodie.xlsx")