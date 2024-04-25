import openpyxl
from peewee import *

db = SqliteDatabase('db/student.db')

class Family(Model):
    user_id_by_yandex = TextField()
    created_at = TextField()
    full_name = TextField()
    classes = TextField()
    birth_date = TextField()
    phone_number = TextField()
    email = TextField()
    birth_place = TextField()
    registration_place = TextField()
    living_place = TextField()
    passport_data = TextField()
    snils = TextField()
    inn = TextField()
    mother_full_name = TextField()
    mother_workplace = TextField()
    mother_phone_number = TextField()
    father_full_name = TextField()
    father_workplace = TextField()
    father_phone_number = TextField()

    class Meta:
        database = db


db.create_tables([Family])

book = openpyxl.load_workbook(r"C:\Users\Anmori\Downloads\Telegram Desktop\Сведения учащихся 2023-2024 на 20.01.24.xlsx", read_only=True)
book.active = 0
sheet = book.active

# print(sheet[10][1].value)
for row in range(1, sheet.max_row):
    # for line in range(sheet.max_row):
        # pass
    Family.create(user_id_by_yandex = str(sheet[row][0].value),
    created_at = str(sheet[row][1].value),
    full_name = str(sheet[row][2].value),
    classes = str(sheet[row][3].value),
    birth_date = str(sheet[row][4].value),
    phone_number = str(sheet[row][5].value),
    email = str(sheet[row][6].value),
    birth_place = str(sheet[row][7].value),
    registration_place = str(sheet[row][8].value),
    living_place = str(sheet[row][9].value),
    passport_data = str(sheet[row][10].value),
    snils = str("".join([i for i in str(sheet[row][11].value) if i.isdigit()])),
    inn = str(sheet[row][12].value),
    mother_full_name = str(sheet[row][13].value),
    mother_workplace = str(sheet[row][14].value),
    mother_phone_number = str(sheet[row][15].value),
    father_full_name = str(sheet[row][16].value),
    father_workplace = str(sheet[row][17].value),
    father_phone_number = str(sheet[row][18].value))